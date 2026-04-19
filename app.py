"""
NextSupply — Banco de Dados de Oportunidades Públicas
Extração precisa dos PDFs Petronect · Dashboard · Export Excel
"""

import re, zipfile
from io import BytesIO
from pathlib import Path
from collections import defaultdict

import pandas as pd
import pdfplumber
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────────────────────────────────────────
# CAMINHOS
# ─────────────────────────────────────────────────────────────────────────────
BASE_DIR     = Path(__file__).parent
ZIPS_DIR     = BASE_DIR / "data" / "zips"
PIPEFY_PATH  = BASE_DIR / "data" / "pipefy_latest.xlsx"
ANALISE_PATH = BASE_DIR / "data" / "analise_precos.xlsx"
GERAL_PATH   = BASE_DIR / "data" / "planilha_geral.xlsx"
ZIPS_DIR.mkdir(parents=True, exist_ok=True)

# ─────────────────────────────────────────────────────────────────────────────
# CONSTANTES
# ─────────────────────────────────────────────────────────────────────────────
FASES_LANCADAS   = {"Enviado pro Trello","LANÇADO","PEDIDO DE COTAÇÃO ENVIADO","COTADAS TOTALMENTE"}
FASES_DECLINADAS = {"Declinadas"}
FASES_ADIADAS    = {"ADIADAS/RELANÇAR"}
FASES_ANDAMENTO  = {"RECEBIDAS"}

ESCOPOS = {
    "Válvula":        ["válvula","valve","globo","gaveta","borboleta","esfera","retenção","alívio"],
    "Bomba":          ["bomba","pump","centrífuga","submersível"],
    "Motor":          ["motor elétrico","motor trifásico","motor monofásico"],
    "Transmissor":    ["transmissor","transmitter","sensor de","medidor de","transdutor"],
    "Filtro":         ["filtro","filter","coalescer","separador de"],
    "Elétrico":       ["disjuntor","chave seccion","quadro elétrico","painel elétrico","cabo elétrico"],
    "Instrumentação": ["instrumento","controlador","indicador","manômetro","pressostato"],
    "Mangueira":      ["mangueira","hose","flexível","tubing"],
    "Rolamento":      ["rolamento","bearing","mancal"],
    "Compressor":     ["compressor"],
    "Atuador":        ["atuador","actuator"],
    "Iluminação":     ["luminária","lanterna","refletor","led","lâmpada"],
    "Câmera":         ["câmera","camera","cftv","vigilância"],
    "Incêndio":       ["incêndio","extintor","sprinkler","hidrante"],
    "Guindaste":      ["guindaste","crane","içamento","talha","caixa redutora","roldana","retentor"],
    "HVAC":           ["hvac","climatizador","ar condicionado","ventilação"],
    "Serviço":        ["serviço","manutenção","inspeção","reparo"],
    "EPI":            ["capacete","luva","botina","óculos protet"],
    "Cabo":           ["cabo de","cabo elétrico","cabo de dados","cabo de rede"],
}

st.set_page_config(page_title="NextSupply | OPS Database", page_icon="🔧", layout="wide")

# ─────────────────────────────────────────────────────────────────────────────
# CSS
# ─────────────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
:root{--orange:#E8630A;--bg:#06111f;--surface:#0c1d30;--surface2:#112540;--text:#e8eef6;--muted:#7a95b0;}
.stApp{background:radial-gradient(ellipse 60% 40% at 0% 0%,rgba(232,99,10,.10),transparent),linear-gradient(180deg,#060f1c 0%,#040b16 100%);color:var(--text);}
.block-container{max-width:1400px;padding-top:1.5rem;}
[data-testid="stHeader"]{background:transparent;}
[data-testid="stSidebar"]{background:#0c1d30 !important;border-right:1px solid rgba(232,99,10,.2)!important;}
[data-testid="stSidebar"] *{color:#e8eef6 !important;}
.stTabs [data-baseweb="tab-list"]{background:transparent;border-bottom:1px solid rgba(232,99,10,.2);gap:.4rem;}
.stTabs [data-baseweb="tab"]{background:rgba(12,29,48,.6);border:1px solid rgba(232,99,10,.15);border-radius:8px 8px 0 0;color:#7a95b0!important;padding:.5rem 1.2rem;font-size:.9rem;}
.stTabs [aria-selected="true"]{background:rgba(232,99,10,.12)!important;border-color:rgba(232,99,10,.4)!important;color:#E8630A!important;}
[data-testid="stDataFrame"]{border:1px solid rgba(232,99,10,.2);border-radius:10px;}
.metric-card{background:linear-gradient(135deg,#0c1d30,#112540);border:1px solid rgba(232,99,10,.22);border-radius:16px;padding:1.5rem 1rem;text-align:center;height:100%;}
.metric-card .val{font-size:2.6rem;font-weight:800;line-height:1;}
.metric-card .pct{font-size:.72rem;margin-top:4px;font-weight:600;}
.metric-card .lbl{font-size:.72rem;color:#7a95b0;margin-top:.5rem;text-transform:uppercase;letter-spacing:.08em;}
.metric-card .sub{font-size:.8rem;color:#a0b4c8;margin-top:.25rem;}
.info-box{background:rgba(232,99,10,.07);border:1px solid rgba(232,99,10,.2);border-radius:10px;padding:.85rem 1.1rem;margin-bottom:1rem;font-size:.87rem;color:#c8d8e8;line-height:1.6;}
.warn-box{background:rgba(251,191,36,.07);border:1px solid rgba(251,191,36,.25);border-radius:10px;padding:.85rem 1.1rem;margin-bottom:.8rem;font-size:.87rem;color:#fde68a;}
.ok-box{background:rgba(34,197,94,.07);border:1px solid rgba(34,197,94,.25);border-radius:10px;padding:.85rem 1.1rem;margin-bottom:.8rem;font-size:.87rem;color:#bbf7d0;}
.sec{font-size:1rem;font-weight:700;color:#E8630A;border-bottom:1px solid rgba(232,99,10,.18);padding-bottom:.4rem;margin:1.4rem 0 .9rem;}
.stButton>button{background:linear-gradient(135deg,#E8630A,#c4510a)!important;color:white!important;border:none!important;border-radius:8px!important;font-weight:600!important;}
.stDownloadButton>button{background:linear-gradient(135deg,#1a3a5c,#0f2540)!important;color:#60a5fa!important;border:1px solid rgba(96,165,250,.3)!important;border-radius:8px!important;font-weight:600!important;}
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────────────
# EXTRAÇÃO DOS PDFs — formato Petronect real
# ─────────────────────────────────────────────────────────────────────────────

def limpar_texto(t: str) -> str:
    t = re.sub(r'\u00ad', '', t)
    t = re.sub(r'Pág:\s*\d+/\d+', '', t)
    t = re.sub(r'Resumo extraído por.*', '', t)
    return re.sub(r' {2,}', ' ', t).strip()

def extrair_numero_op(texto: str) -> str:
    """Extrai o número de 10 dígitos começando com 7."""
    m = re.search(r'ID da Oportunidade\s+(7\d{9})', texto)
    if m: return m.group(1)
    m = re.search(r'\b(7\d{9})\b', texto)
    return m.group(1) if m else None

def extrair_nome_op(texto: str) -> str:
    """Extrai o nome/título da oportunidade."""
    m = re.search(r'Nome da Oportunidade\s+(.+?)(?:\n|Data da publicação)', texto, re.DOTALL)
    if m:
        return m.group(1).strip()[:120]
    # fallback: pega da segunda linha do resumo
    m = re.search(r'Resumo da Oportunidade\s*\n(.+)', texto)
    if m:
        return m.group(1).strip()[:120]
    return ""

def extrair_prazo(texto: str) -> str:
    """Extrai 'Fim do período de cotação' no formato dd.mm.yyyy."""
    m = re.search(r'Fim do per[íi]odo de cota[çc][ãa]o\s+(\d{2}\.\d{2}\.\d{4})', texto)
    if m: return m.group(1).replace('.', '/')
    # fallback: qualquer data dd.mm.yyyy
    m = re.search(r'\b(\d{2}\.\d{2}\.\d{4})\b', texto)
    return m.group(1).replace('.', '/') if m else ""

def extrair_tipo(texto: str) -> str:
    """Extrai tipo de oportunidade."""
    m = re.search(r'Tipo de Oportunidade\s+(.+?)(?:\n|Crit)', texto)
    if m: return m.group(1).strip()
    return ""

def extrair_local(texto: str) -> str:
    """Extrai local de entrega."""
    m = re.search(r'Local de Entrega\s+(.+?)(?:\n|Informa)', texto)
    if m: return m.group(1).strip()
    return ""

def extrair_fabricantes_e_pns(texto: str) -> tuple:
    """
    Extrai fabricantes e Part Numbers a partir de:
    - 'Tp: FABRICANTE PARTNUMBER' nas descrições longas
    - 'FABRICANTE: NOME' explícito
    Retorna (lista_fabricantes, lista_pns)
    """
    fabricantes = []
    pns = []

    # Padrão principal: "Tp: TEXTO" — primeiro token = fabricante, resto = PN
    for m in re.finditer(r'Tp:\s+([A-Z][A-Z0-9\-/&\. ]{2,})', texto):
        # Remove tudo após "---" (separadores de referências extras)
        trecho = re.split(r'\s*-{3,}\s*', m.group(1))[0].strip()
        partes = trecho.split()
        if partes:
            fab = partes[0].rstrip('/')
            pn  = ' '.join(partes[1:]) if len(partes) > 1 else ""
            if fab not in fabricantes:
                fabricantes.append(fab)
            if pn and pn not in pns:
                pns.append(pn)

    # Padrão explícito: "FABRICANTE: NOME"
    for m in re.finditer(r'FABRICANTE:\s+([A-Z][A-Z0-9 &\-\.]+)', texto):
        fab = m.group(1).strip()
        if fab and fab not in fabricantes:
            fabricantes.append(fab)

    return fabricantes, pns

def extrair_escopo(nome: str, texto: str) -> str:
    """Classifica o escopo com base no nome da OP e texto."""
    combined = (nome + " " + texto).lower()
    encontrados = [esc for esc, kws in ESCOPOS.items() if any(kw in combined for kw in kws)]
    return ", ".join(encontrados) if encontrados else "Outros"

def classificar_fase(fase: str) -> str:
    if fase in FASES_LANCADAS:   return "Lançada"
    if fase in FASES_DECLINADAS: return "Declinada"
    if fase in FASES_ADIADAS:    return "Adiada"
    if fase in FASES_ANDAMENTO:  return "Em andamento"
    return "Outro"

# ─────────────────────────────────────────────────────────────────────────────
# CARREGAMENTO DE DADOS
# ─────────────────────────────────────────────────────────────────────────────

@st.cache_data(show_spinner=False)
def carregar_todos_zips():
    arquivos_zip = sorted(ZIPS_DIR.glob("*.zip"))
    if not arquivos_zip:
        return pd.DataFrame(columns=["op","nome","tipo","local","prazo","fabricantes","part_numbers","escopo","arquivo_zip"])

    ops = []
    for zip_path in arquivos_zip:
        try:
            with zipfile.ZipFile(zip_path) as zf:
                pdfs = [n for n in zf.namelist() if n.lower().endswith('.pdf')]
                for pdf_name in pdfs:
                    try:
                        with zf.open(pdf_name) as pf:
                            pdf_bytes = pf.read()
                        with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
                            texto_raw = "\n".join(p.extract_text() or "" for p in pdf.pages)
                        texto = limpar_texto(texto_raw)
                        op = extrair_numero_op(texto)
                        if not op:
                            continue
                        nome   = extrair_nome_op(texto)
                        prazo  = extrair_prazo(texto)
                        tipo   = extrair_tipo(texto)
                        local  = extrair_local(texto)
                        fabs, pns = extrair_fabricantes_e_pns(texto)
                        escopo = extrair_escopo(nome, texto)
                        ops.append({
                            "op":           op,
                            "nome":         nome,
                            "tipo":         tipo,
                            "local":        local,
                            "prazo":        prazo,
                            "fabricantes":  ", ".join(fabs) if fabs else "Não identificado",
                            "part_numbers": ", ".join(pns) if pns else "",
                            "escopo":       escopo,
                            "arquivo_zip":  zip_path.name,
                        })
                    except Exception:
                        pass
        except Exception:
            pass

    if not ops:
        return pd.DataFrame(columns=["op","nome","tipo","local","prazo","fabricantes","part_numbers","escopo","arquivo_zip"])
    return pd.DataFrame(ops).drop_duplicates(subset=["op"])


@st.cache_data(show_spinner=False)
def carregar_pipefy():
    if not PIPEFY_PATH.exists():
        return None
    df = pd.read_excel(PIPEFY_PATH)
    df.columns = df.columns.str.strip()
    col_titulo = "Titulo" if "Titulo" in df.columns else "Título"
    df["op"]        = df[col_titulo].astype(str).str.strip()
    df["fase"]      = df["Fase atual"].astype(str).str.strip()
    df["etiquetas"] = df["Etiquetas"].fillna("").astype(str)
    df["criado_em"] = pd.to_datetime(df["Criado em"], errors="coerce")
    return df[["op","fase","etiquetas","criado_em","Criador"]]


@st.cache_data(show_spinner=False)
def carregar_analise():
    if not ANALISE_PATH.exists():
        return None
    xl = pd.ExcelFile(ANALISE_PATH)
    dados    = pd.read_excel(ANALISE_PATH, sheet_name="DADOS")
    lancados = pd.read_excel(ANALISE_PATH, sheet_name="LANÇADOS") if "LANÇADOS" in xl.sheet_names else pd.DataFrame()
    dados.columns = dados.columns.str.strip()
    return dados, lancados


@st.cache_data(show_spinner=False)
def carregar_geral():
    if not GERAL_PATH.exists():
        return None
    df = pd.read_excel(GERAL_PATH, sheet_name="Comercial")
    df.columns = df.columns.str.strip()
    return df


def cruzar(df_zips, df_pipefy):
    merged = df_zips.merge(df_pipefy, on="op", how="left")
    merged["fase"]      = merged["fase"].fillna("Ignorada")
    merged["etiquetas"] = merged["etiquetas"].fillna("")
    merged["status"]    = merged["fase"].apply(lambda f: "Ignorada" if f == "Ignorada" else classificar_fase(f))
    return merged

# ─────────────────────────────────────────────────────────────────────────────
# EXPORT EXCEL FORMATADO
# ─────────────────────────────────────────────────────────────────────────────

COR_STATUS = {
    "Lançada":      "FF4ADE80",
    "Declinada":    "FFF87171",
    "Ignorada":     "FF9CA3AF",
    "Adiada":       "FFFBBF24",
    "Em andamento": "FF60A5FA",
    "Outro":        "FFA0B4C8",
}

def gerar_excel_bonito(df: pd.DataFrame) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "OPS Comparativo"

    # Cores
    HEADER_FILL  = PatternFill("solid", fgColor="FF0C1D30")
    ORANGE_FILL  = PatternFill("solid", fgColor="FFE8630A")
    ALT_FILL     = PatternFill("solid", fgColor="FF0F2030")
    thin = Side(style="thin", color="FF1E3A52")
    borda = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Título
    ws.merge_cells("A1:I1")
    ws["A1"] = "NextSupply — Banco de Oportunidades Públicas"
    ws["A1"].font = Font(bold=True, size=14, color="FFE8630A")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Subtítulo
    ws.merge_cells("A2:I2")
    ws["A2"] = f"Exportado em {pd.Timestamp.now().strftime('%d/%m/%Y %H:%M')} · Total: {len(df)} oportunidades"
    ws["A2"].font = Font(size=10, color="FF7A95B0")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 18

    # Linha em branco
    ws.row_dimensions[3].height = 8

    # Cabeçalhos das colunas
    colunas = [
        ("Nº OP",          "Número da Oportunidade no Petronect", 14),
        ("Nome / Descrição","Nome completo da oportunidade", 45),
        ("Status",         "Situação no Pipefy: Lançada / Declinada / Ignorada / Adiada", 14),
        ("Fabricante(s)",  "Fabricante(s) identificado(s) no PDF via campo Tp: ou FABRICANTE:", 22),
        ("Part Number(s)", "Part Number(s) extraído(s) do campo Tp: nas descrições dos itens", 28),
        ("Escopo",         "Categoria do material baseada nas descrições dos itens", 28),
        ("Prazo",          "Data limite (Fim do período de cotação)", 12),
        ("Local Entrega",  "Local de entrega indicado na oportunidade", 22),
        ("Arquivo ZIP",    "Nome do arquivo ZIP de origem", 38),
    ]

    for col_idx, (header, _, width) in enumerate(colunas, 1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.fill = PatternFill("solid", fgColor="FFE8630A")
        cell.font = Font(bold=True, size=10, color="FFFFFFFF")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = borda
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[4].height = 24

    # Linha de descrições das colunas
    for col_idx, (_, desc, _) in enumerate(colunas, 1):
        cell = ws.cell(row=5, column=col_idx, value=desc)
        cell.fill = PatternFill("solid", fgColor="FF0F2030")
        cell.font = Font(italic=True, size=8, color="FF7A95B0")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = borda
    ws.row_dimensions[5].height = 30

    # Mapeamento df → colunas Excel
    col_map = ["op","nome","status","fabricantes","part_numbers","escopo","prazo","local","arquivo_zip"]

    # Dados
    for row_idx, (_, row) in enumerate(df.iterrows()):
        excel_row = row_idx + 6
        fill = ALT_FILL if row_idx % 2 == 0 else PatternFill("solid", fgColor="FF0A1828")

        for col_idx, col_key in enumerate(col_map, 1):
            val = row.get(col_key, "")
            cell = ws.cell(row=excel_row, column=col_idx, value=str(val) if val else "")
            cell.fill = fill
            cell.font = Font(size=9, color="FFE8EEF6")
            cell.alignment = Alignment(vertical="center", wrap_text=(col_idx in [2,4,5,6]))
            cell.border = borda

            # Colorir célula de status
            if col_idx == 3:
                cor = COR_STATUS.get(str(val), "FFA0B4C8")
                cell.fill = PatternFill("solid", fgColor=cor)
                cell.font = Font(bold=True, size=9, color="FF06111F")
                cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.row_dimensions[excel_row].height = 18

    # Congelar cabeçalhos
    ws.freeze_panes = "A6"

    # Aba de Resumo
    ws2 = wb.create_sheet("Resumo")
    ws2["A1"] = "Resumo por Status"
    ws2["A1"].font = Font(bold=True, size=13, color="FFE8630A")

    ws2.append([])
    ws2.append(["Status", "Quantidade", "% do Total"])
    for cell in ws2[3]:
        cell.font = Font(bold=True, color="FFFFFFFF")
        cell.fill = PatternFill("solid", fgColor="FFE8630A")
        cell.alignment = Alignment(horizontal="center")

    total = len(df)
    for status, count in df["status"].value_counts().items():
        pct = f"{count/total*100:.1f}%"
        row_data = [status, count, pct]
        ws2.append(row_data)
        cor = COR_STATUS.get(status, "FFA0B4C8")
        ws2.cell(row=ws2.max_row, column=1).fill = PatternFill("solid", fgColor=cor)
        ws2.cell(row=ws2.max_row, column=1).font = Font(bold=True, color="FF06111F")

    ws2["A1"].font = Font(bold=True, size=13, color="FFE8630A")
    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 14

    ws2.append([])
    ws2.append(["Por Escopo/Categoria", "", ""])
    ws2.cell(ws2.max_row, 1).font = Font(bold=True, size=11, color="FFE8630A")

    ws2.append(["Categoria", "Total", "Lançadas"])
    for cell in ws2[ws2.max_row]:
        cell.font = Font(bold=True, color="FFFFFFFF")
        cell.fill = PatternFill("solid", fgColor="FF0C1D30")

    # contagem por escopo
    escopo_contagem = defaultdict(lambda: defaultdict(int))
    for _, row in df.iterrows():
        for tag in str(row.get("escopo","")).split(","):
            t = tag.strip()
            if t:
                escopo_contagem[t][row["status"]] += 1
    for cat, sd in sorted(escopo_contagem.items(), key=lambda x: -sum(x[1].values())):
        ws2.append([cat, sum(sd.values()), sd.get("Lançada", 0)])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()

# ─────────────────────────────────────────────────────────────────────────────
# CABEÇALHO
# ─────────────────────────────────────────────────────────────────────────────
c1, c2 = st.columns([1, 9])
with c1:
    st.markdown("<div style='font-size:3rem;margin-top:.3rem'>🔧</div>", unsafe_allow_html=True)
with c2:
    st.markdown("# NextSupply")
    st.markdown("<span style='color:#7a95b0;font-size:.9rem'>Banco de Oportunidades Públicas · Comparativo ZIP vs Pipefy · Histórico de Cotações</span>", unsafe_allow_html=True)
st.markdown("---")

# ─────────────────────────────────────────────────────────────────────────────
# SIDEBAR
# ─────────────────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("## Status dos Dados")
    n_zips = len(list(ZIPS_DIR.glob("*.zip")))
    st.markdown(f"**ZIPs:** {'✅ ' + str(n_zips) + ' arquivo(s)' if n_zips else '❌ Nenhum em data/zips/'}")
    if n_zips:
        with st.expander(f"Ver {n_zips} ZIPs"):
            for z in sorted(ZIPS_DIR.glob("*.zip")):
                st.caption(z.name)
    st.markdown("---")

    for label, path, key in [
        ("Pipefy (Relatório)", PIPEFY_PATH, "up_pipefy"),
        ("Análise de Preços",  ANALISE_PATH, "up_analise"),
        ("Planilha Geral",     GERAL_PATH,   "up_geral"),
    ]:
        st.markdown(f"**{label}**")
        if path.exists():
            st.markdown(f"✅ {path.name}")
        else:
            st.markdown("❌ Não encontrado")
        arq = st.file_uploader(f"Atualizar {label}", type=["xlsx"], key=key, label_visibility="collapsed")
        if arq:
            path.parent.mkdir(parents=True, exist_ok=True)
            path.write_bytes(arq.read())
            st.cache_data.clear()
            st.success("Atualizado!")
            st.rerun()
        st.markdown("---")

    st.caption("Novos ZIPs → copie para data/zips/ e faça push no GitHub Desktop.")

# ─────────────────────────────────────────────────────────────────────────────
# CARREGAR DADOS
# ─────────────────────────────────────────────────────────────────────────────
with st.spinner("Lendo ZIPs e extraindo dados dos PDFs…"):
    df_zips = carregar_todos_zips()
df_pipefy_data = carregar_pipefy()

# ─────────────────────────────────────────────────────────────────────────────
# ABAS
# ─────────────────────────────────────────────────────────────────────────────
tab1, tab2, tab3 = st.tabs(["📊  Dashboard & Comparativo", "🔍  Histórico de Preços", "📄  Consulta da Sophia"])

# ══════════════════════════════════════════════════════════════════════════════
# TAB 1 — DASHBOARD
# ══════════════════════════════════════════════════════════════════════════════
with tab1:
    if df_zips.empty:
        st.markdown("<div class='warn-box'>⚠️ Nenhum ZIP em <code>data/zips/</code>. Adicione via GitHub Desktop.</div>", unsafe_allow_html=True)
    elif df_pipefy_data is None:
        st.markdown("<div class='warn-box'>⚠️ Relatório Pipefy não encontrado. Faça upload na barra lateral.</div>", unsafe_allow_html=True)
    else:
        df = cruzar(df_zips, df_pipefy_data)
        contagem = df["status"].value_counts()
        total = len(df)

        # ── Métricas principais ──
        st.markdown("<div class='sec'>📈 Visão Geral</div>", unsafe_allow_html=True)
        metricas = [
            ("Total Recebidas", total,                              "#E8630A", "Todas as OPs nos ZIPs"),
            ("Lançadas",        contagem.get("Lançada", 0),        "#4ade80", "Enviadas pro Trello"),
            ("Declinadas",      contagem.get("Declinada", 0),      "#f87171", "Trabalhadas e recusadas"),
            ("Ignoradas",       contagem.get("Ignorada", 0),       "#9ca3af", "Recebidas, não trabalhadas"),
            ("Adiadas",         contagem.get("Adiada", 0),         "#fbbf24", "Para relançar"),
            ("Em andamento",    contagem.get("Em andamento", 0),   "#60a5fa", "No Pipefy agora"),
        ]
        cols = st.columns(6)
        for col, (lbl, val, cor, sub) in zip(cols, metricas):
            pct = f"{val/total*100:.0f}%" if total and lbl != "Total Recebidas" else ""
            col.markdown(f"""
            <div class='metric-card'>
                <div class='val' style='color:{cor}'>{val}</div>
                <div class='pct' style='color:{cor}'>{pct}</div>
                <div class='lbl'>{lbl}</div>
                <div class='sub'>{sub}</div>
            </div>""", unsafe_allow_html=True)

        st.markdown("<br>", unsafe_allow_html=True)

        # ── Gráfico por Categoria ──
        st.markdown("<div class='sec'>🏷️ Por Categoria de Material</div>", unsafe_allow_html=True)

        escopo_status = defaultdict(lambda: defaultdict(int))
        for _, row in df.iterrows():
            for tag in str(row.get("escopo","")).split(","):
                t = tag.strip()
                if t and t != "nan":
                    escopo_status[t][row["status"]] += 1

        linhas_cat = []
        for tag, sd in escopo_status.items():
            tot = sum(sd.values())
            tx  = f"{sd.get('Lançada',0)/tot*100:.0f}%" if tot else "—"
            linhas_cat.append({
                "Categoria":        tag,
                "✅ Lançadas":      sd.get("Lançada", 0),
                "❌ Declinadas":    sd.get("Declinada", 0),
                "⚪ Ignoradas":     sd.get("Ignorada", 0),
                "⏳ Adiadas":      sd.get("Adiada", 0),
                "🔄 Em andamento": sd.get("Em andamento", 0),
                "Total":            tot,
                "Taxa Lançamento":  tx,
            })
        df_cat = pd.DataFrame(linhas_cat).sort_values("Total", ascending=False)

        col_chart, col_table = st.columns([1, 2])
        with col_chart:
            chart_data = df_cat.set_index("Categoria")[["✅ Lançadas","❌ Declinadas","⚪ Ignoradas","⏳ Adiadas"]].head(10)
            st.bar_chart(chart_data, height=380, color=["#4ade80","#f87171","#9ca3af","#fbbf24"])
        with col_table:
            st.dataframe(df_cat, use_container_width=True, hide_index=True, height=380)

        # ── Filtros ──
        st.markdown("<div class='sec'>🔎 Filtros e Detalhamento</div>", unsafe_allow_html=True)
        f1, f2, f3, f4 = st.columns(4)
        filtro_status   = f1.selectbox("Status", ["Todos"] + sorted(df["status"].unique()))
        todas_etiquetas = set()
        for e in df["etiquetas"].dropna():
            for tag in str(e).split(","):
                t = tag.strip()
                if t and t.lower() != "nan":
                    todas_etiquetas.add(t)
        filtro_etiqueta = f2.selectbox("Etiqueta Pipefy", ["Todas"] + sorted(todas_etiquetas))

        todos_escopos = set()
        for e in df["escopo"].dropna():
            for t in str(e).split(","):
                t = t.strip()
                if t: todos_escopos.add(t)
        filtro_escopo = f3.selectbox("Categoria", ["Todas"] + sorted(todos_escopos))
        busca_op      = f4.text_input("Buscar nº OP ou fabricante")

        df_f = df.copy()
        if filtro_status   != "Todos":   df_f = df_f[df_f["status"] == filtro_status]
        if filtro_etiqueta != "Todas":   df_f = df_f[df_f["etiquetas"].str.contains(filtro_etiqueta, na=False)]
        if filtro_escopo   != "Todas":   df_f = df_f[df_f["escopo"].str.contains(filtro_escopo, na=False)]
        if busca_op:
            mask = (df_f["op"].str.contains(busca_op, case=False, na=False) |
                    df_f["fabricantes"].str.contains(busca_op, case=True, na=False) |
                    df_f["nome"].str.contains(busca_op, case=False, na=False))
            df_f = df_f[mask]

        st.markdown(f"<div style='color:#7a95b0;font-size:.85rem;margin-bottom:.5rem'>{len(df_f)} registros encontrados</div>", unsafe_allow_html=True)

        # Tabela detalhada
        colunas_show = [c for c in ["op","nome","status","fabricantes","part_numbers","escopo","prazo","local","etiquetas","arquivo_zip"] if c in df_f.columns]
        df_show = df_f[colunas_show].rename(columns={
            "op":"Nº OP","nome":"Nome da Oportunidade","status":"Status",
            "fabricantes":"Fabricante(s)","part_numbers":"Part Number(s)",
            "escopo":"Categoria","prazo":"Prazo","local":"Local Entrega",
            "etiquetas":"Etiquetas Pipefy","arquivo_zip":"Arquivo ZIP"
        })
        st.dataframe(df_show, use_container_width=True, hide_index=True, height=420)

        # Exportar
        col_exp1, col_exp2, _ = st.columns([1, 1, 3])
        xlsx_bytes = gerar_excel_bonito(df_f.rename(columns={
            "op":"op","nome":"nome","status":"status","fabricantes":"fabricantes",
            "part_numbers":"part_numbers","escopo":"escopo","prazo":"prazo",
            "local":"local","arquivo_zip":"arquivo_zip"
        }))
        col_exp1.download_button("⬇️ Exportar Excel (.xlsx)", data=xlsx_bytes,
            file_name="nextsupply_ops.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        csv_bytes = df_show.to_csv(index=False).encode("utf-8")
        col_exp2.download_button("⬇️ Exportar CSV", data=csv_bytes,
            file_name="nextsupply_ops.csv", mime="text/csv")

# ══════════════════════════════════════════════════════════════════════════════
# TAB 2 — HISTÓRICO DE PREÇOS
# ══════════════════════════════════════════════════════════════════════════════
with tab2:
    st.markdown("<div class='sec'>🔍 Busca no Histórico de Cotações</div>", unsafe_allow_html=True)
    st.markdown("<div class='info-box'>Busca na <strong>Análise de Preços</strong> e na <strong>Planilha Geral de Dispensas</strong> (uso consultivo). Digite Part Number, Fabricante ou Nº OP.</div>", unsafe_allow_html=True)

    analise = carregar_analise()
    geral   = carregar_geral()

    if analise is None and geral is None:
        st.info("Nenhum arquivo de histórico. Faça upload na barra lateral.")
    else:
        c_b, c_t = st.columns([3, 1])
        termo      = c_b.text_input("Part Number, Fabricante ou Nº OP", placeholder="Ex: TC3184N, SEATRAX, 7004575543")
        tipo_busca = c_t.selectbox("Fonte", ["Ambos", "Análise de Preços", "Planilha Geral"])

        if termo:
            t = termo.lower()
            if analise and tipo_busca in ("Ambos", "Análise de Preços"):
                dados, _ = analise
                st.markdown("#### Análise de Preços")
                cols_b = [c for c in ["OP","PART NUMBER","FABRICANTE"] if c in dados.columns]
                mask   = dados[cols_b].apply(lambda col: col.astype(str).str.lower().str.contains(t, na=False)).any(axis=1)
                res    = dados[mask]
                if not res.empty:
                    show = [c for c in ["DATA","OP","ITEM","PART NUMBER","FABRICANTE","NOSSO PREÇO UNIT.","NOSSO PREÇO TOTAL","NOSSO PRAZO","Resultado Esperado"] if c in res.columns]
                    st.dataframe(res[show], use_container_width=True, hide_index=True)
                    st.caption(f"{len(res)} registros encontrados na Análise de Preços")
                else:
                    st.info("Nenhum resultado na Análise de Preços.")

            if geral is not None and tipo_busca in ("Ambos", "Planilha Geral"):
                st.markdown("#### Planilha Geral — Dispensas *(uso consultivo)*")
                cols_g = [c for c in ["Número da OP","PN","FABRICANTE","DESCRIÇÃO BREVE"] if c in geral.columns]
                mask_g = geral[cols_g].apply(lambda col: col.astype(str).str.lower().str.contains(t, na=False)).any(axis=1)
                res_g  = geral[mask_g]
                if not res_g.empty:
                    show_g = [c for c in ["Número da OP","ITEM","FABRICANTE","PN","DESCRIÇÃO BREVE","VALOR UNITÁRIO","VALOR TOTAL","COTADOR"] if c in res_g.columns]
                    st.dataframe(res_g[show_g], use_container_width=True, hide_index=True)
                    st.caption(f"{len(res_g)} registros (dispensas — uso consultivo)")
                else:
                    st.info("Nenhum resultado na Planilha Geral.")

# ══════════════════════════════════════════════════════════════════════════════
# TAB 3 — MÓDULO SOPHIA
# ══════════════════════════════════════════════════════════════════════════════
with tab3:
    st.markdown("<div class='sec'>📄 Consulta da Sophia</div>", unsafe_allow_html=True)
    st.markdown(
        "<div class='info-box'>"
        "Suba o PDF de uma oportunidade pública e descubra:<br>"
        "✅ Já foi lançada? &nbsp;❌ Foi declinada? &nbsp;⚪ Nunca trabalhamos?<br>"
        "E verifica automaticamente se existe histórico de preço na Análise de Preços."
        "</div>", unsafe_allow_html=True)

    pdf_sophia = st.file_uploader("📎 PDF da oportunidade", type=["pdf"], key="sophia_pdf")

    if pdf_sophia:
        conteudo_pdf = pdf_sophia.read()
        with st.spinner("Lendo PDF…"):
            erro, texto = None, ""
            try:
                with pdfplumber.open(BytesIO(conteudo_pdf)) as pdf:
                    texto = "\n".join(p.extract_text() or "" for p in pdf.pages)
                texto = limpar_texto(texto)
            except Exception as e:
                erro = str(e)

        if erro:
            st.error(f"Erro ao abrir PDF: {erro}")
        elif not texto.strip():
            st.error("Não foi possível extrair texto. PDF pode ser escaneado ou protegido.")
        else:
            op_num   = extrair_numero_op(texto)
            nome     = extrair_nome_op(texto)
            prazo    = extrair_prazo(texto)
            tipo     = extrair_tipo(texto)
            local    = extrair_local(texto)
            fabs, pns = extrair_fabricantes_e_pns(texto)
            escopo   = extrair_escopo(nome, texto)

            # ── Dados identificados ──
            st.markdown("<div class='sec'>📌 Dados identificados no PDF</div>", unsafe_allow_html=True)
            ca, cb, cc, cd = st.columns(4)
            ca.metric("Nº OP",    op_num or "Não identificado")
            cb.metric("Tipo",     tipo or "—")
            cc.metric("Prazo",    prazo or "—")
            cd.metric("Local",    local or "—")

            if nome:
                st.markdown(f"**Nome da Oportunidade:** {nome}")

            col_fab, col_pn, col_esc = st.columns(3)
            with col_fab:
                st.markdown("**Fabricante(s) identificado(s):**")
                if fabs:
                    for f in fabs: st.markdown(f"- `{f}`")
                else:
                    st.caption("Nenhum identificado")
            with col_pn:
                st.markdown("**Part Number(s) identificado(s):**")
                if pns:
                    for p in pns: st.markdown(f"- `{p}`")
                else:
                    st.caption("Nenhum identificado")
            with col_esc:
                st.markdown("**Categoria detectada:**")
                st.markdown(f"`{escopo}`")

            with st.expander("🔍 Ver texto bruto extraído do PDF"):
                st.text(texto[:3000])

            st.markdown("---")

            # ── Status no Pipefy ──
            st.markdown("<div class='sec'>📊 Status no Pipefy</div>", unsafe_allow_html=True)
            if df_pipefy_data is None:
                st.info("Carregue o relatório Pipefy na barra lateral para verificar o status.")
            elif not op_num:
                st.warning("Número de OP não identificado no PDF.")
            else:
                linha = df_pipefy_data[df_pipefy_data["op"] == op_num]
                if not linha.empty:
                    fase = linha.iloc[0]["fase"]
                    etiq = linha.iloc[0]["etiquetas"]
                    cri  = linha.iloc[0]["criado_em"]
                    status = classificar_fase(fase)
                    box_map = {
                        "Lançada":      ("ok-box",   "✅ LANÇADA"),
                        "Declinada":    ("warn-box", "❌ DECLINADA"),
                        "Adiada":       ("warn-box", "⏳ ADIADA"),
                        "Em andamento": ("info-box", "🔄 EM ANDAMENTO"),
                        "Outro":        ("info-box", "ℹ️ NO PIPEFY"),
                    }
                    box_cls, titulo = box_map.get(status, ("info-box","ℹ️ NO PIPEFY"))
                    st.markdown(
                        f"<div class='{box_cls}'><strong>{titulo}</strong><br>"
                        f"Fase: <strong>{fase}</strong> &nbsp;|&nbsp; Etiquetas: {etiq} &nbsp;|&nbsp; "
                        f"Criado em: {str(cri)[:10] if pd.notna(cri) else '-'}</div>",
                        unsafe_allow_html=True)
                else:
                    if not df_zips.empty and op_num in df_zips["op"].values:
                        st.markdown("<div class='warn-box'>⚪ <strong>IGNORADA</strong> — Recebida no ZIP mas nunca trabalhada no Pipefy.</div>", unsafe_allow_html=True)
                    else:
                        st.markdown("<div class='warn-box'>⚪ <strong>NÃO ENCONTRADA</strong> — Não está nos ZIPs nem no Pipefy.</div>", unsafe_allow_html=True)

            # ── Histórico de Preço ──
            st.markdown("<div class='sec'>💰 Histórico na Análise de Preços</div>", unsafe_allow_html=True)
            analise = carregar_analise()
            if analise is None:
                st.info("Carregue a Análise de Preços na barra lateral.")
            else:
                dados, _ = analise
                termos = list(set(filter(None, [op_num] + fabs + pns)))
                termos = [t for t in termos if t != "Não identificado"]
                if not termos:
                    st.warning("Sem termos para buscar (OP, fabricante e PN não identificados).")
                else:
                    resultados = []
                    for termo in termos:
                        cols_b = [c for c in ["OP","PART NUMBER","FABRICANTE"] if c in dados.columns]
                        mask = dados[cols_b].apply(
                            lambda col: col.astype(str).str.upper().str.contains(str(termo).upper(), na=False)
                        ).any(axis=1)
                        resultados.append(dados[mask])
                    df_hist = pd.concat(resultados).drop_duplicates() if resultados else pd.DataFrame()
                    if not df_hist.empty:
                        st.success(f"✅ {len(df_hist)} registros históricos encontrados!")
                        show = [c for c in ["DATA","OP","ITEM","PART NUMBER","FABRICANTE","NOSSO PREÇO UNIT.","NOSSO PRAZO","Resultado Esperado"] if c in df_hist.columns]
                        if not show: show = list(df_hist.columns[:8])
                        st.dataframe(df_hist[show], use_container_width=True, hide_index=True)
                    else:
                        st.info("⚪ Nenhum histórico encontrado. Termos buscados: " + ", ".join(str(t) for t in termos))
